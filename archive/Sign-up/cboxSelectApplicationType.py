from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
from datetime import datetime
from openpyxl import load_workbook #Excel Access
import os
import sys
import time
import importlib

#initiate driver
driver = webdriver.Chrome()
action = webdriver.ActionChains(driver)

#open scripts
wb = load_workbook(r"C:\Scripts\20231018_Truck_Booking_v1 - Build 27.2.xlsx")
sheet = wb.active
TotalRow = sheet.max_row

#def call(driver, TestURL):
# try:
#     driver.get("http://112.199.119.250:82/peza/login")
#     driver.maximize_window()
#     print("Test Passed: Website initiated")
#     #return 1
# except:
#     print("Cannot not open website: Test Fail")
#     #return 0
#
# time.sleep(5)

# try:
#     button = driver.find_element(By.XPATH, '//span[text()="Create an account"]')
#     driver.execute_script("arguments[0].click();", button)
#     print("Sign-up page, Passed!")
#     #return 1
# except NoSuchElementException:
#     print("Sign-up page, failed!")
#     #return 0
#
# time.sleep(5)

# try:
#     btnProceedComInfo = driver.find_element(By.XPATH, '//input[@name="company.registration_type_id" and @value=2]')
#     driver.execute_script("arguments[0].click();", btnProceedComInfo)
#     print("Proceed to Company Info")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0
#
# time.sleep(1)

# try:
#     ProceedCompanyInfo = driver.find_element(By.XPATH, '//button[@id="btn_comp_info"]')
#     driver.execute_script("arguments[0].click();", ProceedCompanyInfo)
#     print("Test Passed: Selected Use Own Truck")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

time.sleep(5)

# try:
#     InputCompanyName = driver.find_element(By.XPATH, '//input[@id="companyName"]')
#     InputCompanyName.send_keys("WHY THE COMPANY NAME D")
#     print("Company NAME INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0
#
# time.sleep(1)

# try:
#     InputCompanyTIN = driver.find_element(By.XPATH, '//input[@id="companyTinNumber"]')
#     InputCompanyTIN.send_keys("000999000003")
#     print("Company TIN INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)
#
# try:
#     InputCompanyRegion = driver.find_element(By.XPATH, '//select[@id="companyRegionId"]')
#     Select(InputCompanyRegion).select_by_index(3)
#     print("Address Region INPUT")
#     #return 1
#
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0
#
# time.sleep(1)

# try:
#     InputCompanyProvince = driver.find_element(By.XPATH, '//select[@id="companyProvinceId"]')
#     Select(InputCompanyProvince).select_by_index(2)
#     print("Address Province INPUT")
#     #return 1
#
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0
#
# time.sleep(1)

# try:
#     InputCompanyCity = driver.find_element(By.XPATH, '//select[@id="companyCityId"]')
#     Select(InputCompanyCity).select_by_index(6)
#     print("Address City INPUT")
#     #return 1
#
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)

# try:
#     InputCompanyBarangay = driver.find_element(By.XPATH, '//select[@id="companyBarangayId"]')
#     Select(InputCompanyBarangay).select_by_index(51)
#     print("Address Barangay INPUT")
#     #return 1
#
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)

# try:
#     InputCompanyStreet = driver.find_element(By.XPATH, '//input[@id="companyAddressLine1"]')
#     InputCompanyStreet.send_keys("WHY THE COMPANY STREET")
#     print("Company Street INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)

# try:
#     InputCompanyMobileNum = driver.find_element(By.XPATH, '//input[@id="companyMobileNumber"]')
#     InputCompanyMobileNum.send_keys("9999999992")
#     print("Company Mobile Number INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0
#
# time.sleep(1)

# try:
#     InputCompanyTelNum = driver.find_element(By.XPATH, '//input[@id="companyTelephoneNumber"]')
#     InputCompanyTelNum.send_keys("289999991")
#     print("Company Telphone Number INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0
#
# time.sleep(1)

# try:
#     InputCompanyEmail = driver.find_element(By.XPATH, '//input[@id="companyEmailAddress"]')
#     InputCompanyEmail.send_keys("rramos.ascent@gmail.com")
#     print("Company Email INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)

# try:
#     ProceedSupportingDocs = driver.find_element(By.XPATH, '//button[@id="btn_supp_docs"]')
#     driver.execute_script("arguments[0].click();", ProceedSupportingDocs)
#     print("Proceed to Supporting Documents")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(5)
#
# try:
#     SupDocsFileLoc = driver.find_element(By.XPATH, '//input[@id="supportingDocumentFilename"]')
#     SupDocsFileLoc.send_keys(r"C:\Users\ASCENT\Downloads\WIMS Proposal - Activity Diagram.pdf")
#     print("Company Supporting Docs File Location INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)
#
# try:
#     SupDocsFileDesc = driver.find_element(By.XPATH, '//input[@id="supportingDocumentDescription"]')
#     SupDocsFileDesc.send_keys("WIMS Proposal - Activity Diagram pdf file")
#     print("Company Supporting Docs description INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)
#
# try:
#     SupDocsValidity = driver.find_element(By.XPATH, '//html//body//form//div//div//div//div//div//div[2]//div//div//div[2]//div//div[3]//div[2]//div//div[1]//div//div[1]//div[3]//input[1]')
#     #SupDocsValidity.send_keys("12/31/2023")
#     driver.execute_script("arguments[0].value =" + "'12/31/2024'" + ";", SupDocsValidity)
#     #SupDocsValidity.text("12/31/2024")
#     print("Company Supporting Docs validity INPUT ", SupDocsValidity.get_property("value"))
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0
#
# time.sleep(1)

# try:
#     btnAddSuppFile = driver.find_element(By.XPATH, '//button[@id="btnAddSupportingDocument"]')
#     driver.execute_script("arguments[0].click();", btnAddSuppFile)
#     print("Supporting documents added")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)
#
# try:
#     btnProceedContactInfo = driver.find_element(By.XPATH, '//button[@id="btn_contact_info"]')
#     driver.execute_script("arguments[0].click();", btnProceedContactInfo)
#     print("Proceed contact info")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0
#
# time.sleep(5)

# try:
#     ContactLastName = driver.find_element(By.XPATH, '//input[@id="lastname"]')
#     ContactLastName.send_keys("WHY THE LASTNAME")
#     print("Contact Lastname INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)
#
# try:
#     ContactFirstName = driver.find_element(By.XPATH, '//input[@id="firstname"]')
#     ContactFirstName.send_keys("WHY THE FIRSTNAME")
#     print("Contact Firstname INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)
#
# try:
#     ContactMiddleName = driver.find_element(By.XPATH, '//input[@id="middlename"]')
#     ContactMiddleName.send_keys("WHY THE MIDDLENAME")
#     print("Contact Middlename INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0
#
# time.sleep(1)
#
# try:
#     ContactCompPosition = driver.find_element(By.XPATH, '//input[@id="companyPosition"]')
#     ContactCompPosition.send_keys("WHY THE POSITION")
#     print("Contact Position INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0
#
# time.sleep(1)

# try:
#     ContactMobile = driver.find_element(By.XPATH, '//input[@id="mobileNumber"]')
#     ContactMobile.send_keys("9999999991")
#     print("Contact Mobile Number INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)

# try:
#     ContactEMail = driver.find_element(By.XPATH, '//input[@id="email"]')
#     ContactEMail.send_keys("rramos.ascent+at004@gmail.com")
#     print("Contact EMail INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0
#
# time.sleep(1)

# try:
#     ContactUserName = driver.find_element(By.XPATH, '//input[@id="username"]')
#     ContactUserName.send_keys("RUSSELRAMOS0004")
#     print("Contact Username INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

time.sleep(1)


# try:
#     ContactPassword = driver.find_element(By.XPATH, '//input[@id="password"]')
#     ContactPassword.send_keys("RUSSELRAMOS0004")
#     print("Contact Password INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)

# try:
#     ContactPasswordConfirm = driver.find_element(By.XPATH, '//input[@id="confirmPassword"]')
#     ContactPasswordConfirm.send_keys("RUSSELRAMOS0004")
#     print("Contact Password Confirm INPUT")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(1)

# try:
#     btnPrcdDetailsConfirm = driver.find_element(By.XPATH, '//button[@id="btn_details_confirm"]')
#     driver.execute_script("arguments[0].click();", btnPrcdDetailsConfirm)
#     print("Proceed details confirmation")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(5)

# try:
#     btnSignupTermsCond = driver.find_element(By.XPATH, '//button[@class="btn btn-primary btn-submit skip-submit-loading skip-text-change"]')
#     driver.execute_script("arguments[0].click();", btnSignupTermsCond)
#     print("Terms and Condition info")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

# time.sleep(2)

# try:
#     TermsConditionModal: WebElement = driver.find_element(By.XPATH, '//div[@id="termsAndConditionModal"]')
#     print("Terms and Condition info modal")
#     try:
#         btnSubmit = TermsConditionModal.find_element(By.XPATH, '//button[@id="btnSubmit"]')
#         #btnStatus = driver.execute_script("arguments[0].disabled;", btnSubmit)
#         btnStatus = btnSubmit.get_attribute("disabled")
#         print(btnStatus)
#         if btnStatus == "true":
#             try:
#                 cboxTermsConditionAccept = TermsConditionModal.find_element(By.XPATH, '//input[@id="agreeCheckbox"]')
#                 driver.execute_script("arguments[0].click();", cboxTermsConditionAccept)
#                 print("Terms and Condition info modal, acceptance checkbox")
#                 # return 1
#             except NoSuchElementException:
#                 print("Element Not found! Test Fail, 4")
#                 #return 0
#         else:
#             print("Element Not found! Test Fail, 3")
#             # return 0
#     except NoSuchElementException:
#         print("Element Not found! Test Fail, 2")
#         # return 0
# except NoSuchElementException:
#     print("Element Not found! Test Fail, 1")
#     #return 0
#
# time.sleep(1)

# try:
#     TermsConditionModal: WebElement = driver.find_element(By.XPATH, '//div[@id="termsAndConditionModal"]')
#     btnSubmit = TermsConditionModal.find_element(By.XPATH, '//button[@id="btnSubmit"]')
#     driver.execute_script("arguments[0].click();", btnSubmit)
#     print("Terms and Condition Acceptance")
#     time.sleep(5)
#     try:
#         SuccessModal: WebElement = driver.find_element(By.XPATH, '//div[@id="swal2-html-container"]')
#         print("Sign-up Success")
#     except NoSuchElementException:
#         print("Sing-up Failed")
#         # return 0
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

time.sleep(30)

driver.close()