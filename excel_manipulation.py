from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.selenium_manager import SeleniumManager
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
from datetime import datetime
from openpyxl import load_workbook #Excel Access
import os
import sys
import time
import importlib


# initiate driver
driver = webdriver.Chrome()
action = webdriver.ActionChains(driver)

# open scripts
wb = load_workbook(r"C:\Scripts\PEZA CP\20231130_PEZA_CPF_v1 - Sign-up D.xlsx")
sheet = wb.active
TotalRow = sheet.max_row

try:
    for x in range(TotalRow):
        CellRow = 0
        if x > 0:
            CellRow = x + 1
            PreviousCell = sheet.cell(row=x, column=1).value
            CurrentCell = sheet.cell(row=CellRow, column=1).value
            if PreviousCell != CurrentCell:
                if x == 1:
                    print("Starting Test Scenario: ", CurrentCell)
                elif not CurrentCell:
                    print(PreviousCell, "Test Scenario is Completed")
                    break
                else:
                    print(PreviousCell, "Test Scenario is Completed")
                    print("Starting Test Scenario: ", CurrentCell)
            TestExecutionStatus = ''

            TestScriptStep = sheet.cell(row=CellRow, column=2).value
            ScriptLocation = sheet.cell(row=CellRow, column=3).value
            TestScript = sheet.cell(row=CellRow, column=4)
            TestData = sheet.cell(row=CellRow, column=5).value

            sys.path.append(ScriptLocation)

            # Select Script
            TestScriptRun = TestScript.value

            # import module
            module = importlib.import_module(TestScriptRun)
            print(CurrentCell, " ", TestScriptStep, "  executing: ", TestScriptRun, ": ", end='')

            # call function
            if not TestData:
                TestExecutionStatus = module.call(driver)
            else:
                TestExecutionStatus = module.call(driver, TestData)

            # Write Response to Excel file
            sheet.cell(row=CellRow, column=6, value=TestExecutionStatus)
            time.sleep(1)
            if TestExecutionStatus == 1:
                print("Passed")
            else:
                print("Failed")

        else:
            print("Initiating Automated Testing")
            TestStartDateTime: datetime = datetime.now()

    TestScriptRun = 'Saving File and Exit Test Run'
    wb.save(r"C:\Scripts\PEZA CP\20231130_PEZA_CPF_v1 - Sign-up D.xlsx")

    print("Test Started: ", TestStartDateTime)
    print("Test Complete: ", datetime.now())
    time.sleep(5)
    driver.close()
except NoSuchElementException:
    print(TestScriptRun, " Script Error")
    wb.save(r"C:\Scripts\PEZA CP\20231130_PEZA_CPF_v1 - Sign-up D.xlsx")

