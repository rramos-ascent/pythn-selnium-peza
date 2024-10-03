import time
from datetime import datetime
from openpyxl import load_workbook
from FrameWork.config_files.frame_work_base_driver import FrameWorkBaseDriver
from FrameWork.peza.peza_execute_test_cases.ecai_execute_test_cases import ECAIExecuteTestCases
from FrameWork.pkg_utilities.utility_script_general import UtilityPackage


class ECAIExecutionPlanExcelRead(FrameWorkBaseDriver):
    def __init__(self, driver):
        self.driver = driver

    def open_test_case_excel_file(self, open_excel_path_location):
        get_work_book = load_workbook(open_excel_path_location)
        get_active_sheet = get_work_book.active
        get_total_rows = get_active_sheet.max_row
        return get_work_book, get_active_sheet, get_total_rows

    def test_execution_plan_read_excel(self, excel_path_location):
        scripting_utility_tools = UtilityPackage()
        peza_call_test_cases = ECAIExecuteTestCases(self.driver)

        scripting_utility_tools.terminate_first_displayed_line()
        get_open_test_case_excel_file = self.open_test_case_excel_file(excel_path_location)

        work_book = get_open_test_case_excel_file[0]
        sheet = get_open_test_case_excel_file[1]
        total_test_execution_row = get_open_test_case_excel_file[2]

        for x in range(total_test_execution_row):
            CellRow = 0
            TestExecutionStatus = ''
            TestExecutionModuleScript = ()
            if x > 0:
                CellRow = x + 1
                PreviousCell = sheet.cell(row=x, column=1).value
                CurrentCell = sheet.cell(row=CellRow, column=1).value
                TestScriptStep = sheet.cell(row=CellRow, column=2).value
                test_execution_module = sheet.cell(row=CellRow, column=3).value
                test_execution_script = sheet.cell(row=CellRow, column=4).value
                test_execution_data = sheet.cell(row=CellRow, column=5).value
                test_execution_module_script = (test_execution_module, test_execution_script, test_execution_data)

                if PreviousCell != CurrentCell:
                    if x == 1:
                        print("Starting Test Scenario: ", CurrentCell)
                    elif not CurrentCell:
                        print(PreviousCell, "Test Scenario is Completed")
                        break
                    else:
                        print(PreviousCell, "Test Scenario is Completed")
                        print("Starting Test Scenario: ", CurrentCell)

                test_execution_func_return = peza_call_test_cases.function_call_test_cases(test_execution_module_script)
                test_execution_script = test_execution_func_return[1]
                test_execution_status = scripting_utility_tools.get_if_passed_failed(test_execution_func_return[0])

                print(CurrentCell, " ", TestScriptStep, "  executing: ", test_execution_script,  ": ", test_execution_status)

                sheet.cell(row=CellRow, column=6, value=test_execution_status)

                time.sleep(1)

            else:
                print("Initiating Automated Testing")
                test_start_date_time: datetime = datetime.now()

        work_book.save(excel_path_location)

        print("Test Started: ", test_start_date_time)
        print("Test Complete: ", datetime.now())
        time.sleep(5)



