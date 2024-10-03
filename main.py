from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
from datetime import datetime
from openpyxl import load_workbook #Excel Access
import os
import sys
import time
import importlib
from datetime import date
from datetime import datetime
from pathlib import Path

#initiate driver
driver = webdriver.Chrome()
wait = WebDriverWait(driver, 10)
action = webdriver.ActionChains(driver)

#open scripts
#wb = load_workbook(r"C:\Scripts\20231018_Truck_Booking_v1 - Build 27.2.xlsx")
#sheet = wb.active
#TotalRow = sheet.max_row

#def call(driver, TestURL):

JS_DROP_FILE = """
    var target = arguments[0],
        offsetX = arguments[1],
        offsetY = arguments[2],
        document = target.ownerDocument || document,
        window = document.defaultView || window;

    var input = document.createElement('INPUT');
    input.type = 'file';
    input.onchange = function () {
      var rect = target.getBoundingClientRect(),
          x = rect.left + (offsetX || (rect.width >> 1)),
          y = rect.top + (offsetY || (rect.height >> 1)),
          dataTransfer = { files: this.files };

      ['dragenter', 'dragover', 'drop'].forEach(function (name) {
        var evt = document.createEvent('MouseEvent');
        evt.initMouseEvent(name, !0, !0, window, 0, 0, 0, x, y, !1, !1, !1, !1, 0, null);
        evt.dataTransfer = dataTransfer;
        target.dispatchEvent(evt);
      });

      setTimeout(function () { document.body.removeChild(input); }, 25);
    };
    document.body.appendChild(input);
    return input;
"""

try:
    driver.get("http://112.199.119.250:82/peza/login")
    driver.maximize_window()
    print("Test Passed: Website initiated")
    #return 1
except:
    print("Cannot not open website: Test Fail")
    #return 0

try:
    # InputUserName = driver.find_element(By.XPATH, '//input[@id="username"]')
    InputUserName = wait.until(EC.element_to_be_clickable((By.XPATH, '//input[@id="username"]')))
    InputUserName.send_keys("RUSSELRAMOS0005")
    print("Username Input")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    InputPassword = driver.find_element(By.XPATH, '//input[@id="password"]')
    InputPassword.send_keys("RUSSELRAMOS0005")
    print("Password Input")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    ProceedLogin = driver.find_element(By.XPATH, '//button[text()="Sign in"]')
    driver.execute_script("arguments[0].click();", ProceedLogin)
    print("Login Initiated")
    try:
        driver.find_element(By.XPATH, '//html[@class="light-style layout-navbar-fixed layout-menu-fixed layout-footer-fixed"]')
        print("Login Success")
        #return 1
    except NoSuchElementException:
        print("Login failed")
        #return 0
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    menuAppAcc = driver.find_element(By.XPATH, '//*[text()=\"Applications & Accreditations\"]')
    driver.execute_script("arguments[0].click();", menuAppAcc)
    print("Application and Accreditation")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

# For Developer
try:
    menuNewBerms = driver.find_element(By.XPATH, '//a[@class="btn btn-primary float-end ms-2"]')
    driver.execute_script("arguments[0].click();", menuNewBerms)
    print("New BERMS application form")
    try:
        driver.find_element(By.XPATH, '//*[text()=\"Application for Registration - Ecozone Developer\"]')
        print("Application form initiated")
    except NoSuchElementException:
        print("Application form failed to load")
        # return 0
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

# For Enterprise
# try:
#     menuNewBerms = driver.find_element(By.XPATH, '//a[@class="btn btn-primary float-end"]')
#     driver.execute_script("arguments[0].click();", menuNewBerms)
#     print("New BERMS application form")
#     time.sleep(1)
#     try:
#         UnderTakingModal: WebElement = driver.find_element(By.XPATH, '//div[@id="undertakingsModal"]')
#         cboxUnderTaking = UnderTakingModal.find_element(By.XPATH, '//input[@id="acceptApplicantUndertakings"]')
#         driver.execute_script("arguments[0].click();", cboxUnderTaking)
#         print("Agree applicant undertaking")
#         time.sleep(1)
#         try:
#             UnderTakingModal1: WebElement = driver.find_element(By.XPATH, '//div[@id="undertakingsModal"]')
#             btnSubmit_undertaking: WebElement = driver.find_element(By.XPATH, '//button[@id="btnSubmitUndertaking"]')
#             driver.execute_script("arguments[0].click();", btnSubmit_undertaking)
#             print("Proceed applicant undertaking")
#             time.sleep(1)
#             try:
#                 driver.find_element(By.XPATH, '//*[text()=\"Application for Registration - Ecozone Enterprise\"]')
#                 print("Application form initiated")
#                 time.sleep(1)
#             except NoSuchElementException:
#                 print("Application form failed to load")
#             # return 0
#         except NoSuchElementException:
#             print("Application form failed to load")
#             # return 0
#     except NoSuchElementException:
#         print("Modal for undertaking is not found")
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

time.sleep(5)

# For Developer
try:
    rboxPezaDeveloper = driver.find_element(By.XPATH, '//input[@id="economicZoneId4"]')
    driver.execute_script("arguments[0].click();", rboxPezaDeveloper)
    print("Select type of Zone Development")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

# For Enterprise
#
# try:
#     rboxPezaEnterprise = driver.find_element(By.XPATH, '//input[@id="enterpriseTypeId5"]')
#     driver.execute_script("arguments[0].click();", rboxPezaEnterprise)
#     print("Select type of Zone Enterprise")
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0
#
# time.sleep(1)

# For Developer

try:
    InputUserName = driver.find_element(By.XPATH, '//input[@id="proposedEconomicZoneName"]')
    InputUserName.send_keys("WHY THE MANUFACTURING ECONOMIC ZONE")
    print("NAME OF MANUFACTURING ECONOMIC ZONE")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

# For Enterprise as start

try:
    InputUserName = driver.find_element(By.XPATH, '//input[@id="enterpriseName"]')
    InputUserName.send_keys("WHY THE PEZA ENTERPRISE AC")
    print("NAME OF PEZA ENTERPRISE")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    InputPSICsec = driver.find_element(By.XPATH, '//select[@name="activity.psic_id"]')
    Select(InputPSICsec).select_by_index(4)
    print("Select type of PSIC Section")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    InputPSIC_div = driver.find_element(By.XPATH, '//select[@name="activity.psic_division_code_id"]')
    Select(InputPSIC_div).select_by_index(1)
    print("Select type of PSIC Division")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    btn_berms_application_proceed = driver.find_element(By.XPATH, '//form[@id="form_application"]//div[1]//div[5]//div//div//button')
    driver.execute_script("arguments[0].click();", btn_berms_application_proceed)
    print("Move to next page ")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    input_berms_nature_business = driver.find_element(By.XPATH, '//select[@id="industryId"]')
    Select(input_berms_nature_business).select_by_visible_text("Education")
    print("Select Nature of Business")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_representative_firstname = driver.find_element(By.XPATH, '//input[@id="representativeFirstName"]')
    input_berms_representative_firstname.send_keys("WHYENTERPRISEREPFIRSTNAME")
    print("Official Representative First Name")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_representative_firstname = driver.find_element(By.XPATH, '//input[@id="representativeLastName"]')
    input_berms_representative_firstname.send_keys("WHYENTERPRISEREPLASTNAME")
    print("Official Representative Last Name")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_rep_home_region = driver.find_element(By.XPATH, '//select[@id="representativeRegionId"]')
    Select(input_berms_rep_home_region).select_by_visible_text("NATIONAL CAPITAL REGION (NCR)")
    print("Select Home address - Region")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_rep_home_province = driver.find_element(By.XPATH, '//select[@id="representativeProvinceId"]')
    Select(input_berms_rep_home_province).select_by_visible_text("METRO MANILA, FOURTH DISTRICT")
    print("Select Home address - Province")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_rep_home_city = driver.find_element(By.XPATH, '//select[@id="representativeCityId"]')
    Select(input_berms_rep_home_city).select_by_visible_text("CITY OF MAKATI")
    print("Select Home address - City")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_rep_home_barangay = driver.find_element(By.XPATH, '//select[@id="representativeBarangayId"]')
    Select(input_berms_rep_home_barangay).select_by_visible_text("GUADALUPE VIEJO")
    print("Select Home address - Barangay")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_rep_home_street = driver.find_element(By.XPATH, '//input[@id="representativeStreetName"]')
    input_berms_rep_home_street.send_keys("WHYENTERPRISEREPSTREETNAME")
    print("Select Home address - Street")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_rep_contct_mobile = driver.find_element(By.XPATH, '//input[@id="representativeMobileNumber"]')
    input_berms_rep_contct_mobile.send_keys("991000001")
    print("Select Home contact - mobile")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_rep_contct_landline = driver.find_element(By.XPATH, '//input[@id="representativetelephoneNumber"]')
    input_berms_rep_contct_landline.send_keys("8881001")
    print("Select Home contact - landline")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_rep_contct_email = driver.find_element(By.XPATH, '//input[@id="representativeEmailAddress"]')
    input_berms_rep_contct_email.send_keys("ramos.ascent+berms_ent_001@gmail.com")
    print("Select Home contact - email")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    btn_berms_rep_proceed = driver.find_element(By.XPATH, '//form[@id="form_application"]//div[2]//div[3]//div//div//button')
    driver.execute_script("arguments[0].click();", btn_berms_rep_proceed)
    print("Move to next page ")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    input_berms_business_prdctactvty = driver.find_element(By.XPATH, '//input[@id="proposedBusinessActivity"]')
    input_berms_business_prdctactvty.send_keys("WHY THE PRODUCT ACTIVITY")
    print("Business product/activity - Product/Activity")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_business_prdctactvty_description = driver.find_element(By.XPATH, '//textarea[@name="activity.description"]')
    input_berms_business_prdctactvty_description.send_keys("WHY THE PRODUCT ACTIVITY WHY THE PRODUCT ACTIVITY WHY THE PRODUCT ACTIVITY WHY THE PRODUCT ACTIVITY WHY THE PRODUCT ACTIVITY")
    print("Business product/activity - description")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_business_prdctactvty_uses = driver.find_element(By.XPATH, '//input[@id="business_prod_service_application"]')
    input_berms_business_prdctactvty_uses.send_keys("WHY THE PRODUCT ACTIVITY USES")
    print("Business product/activity - uses")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_business_prdctactvty_image = driver.find_element(By.XPATH, "//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[3]//div[1]//div//div[2]//div[2]//div//div")
    file_input = driver.execute_script(JS_DROP_FILE, input_berms_business_prdctactvty_image, 0, 0)
    file_input.send_keys(r"C:\Users\ASCENT\Downloads\viber_image_2024-01-02_15-34-22-566.png")
    print("Business product/activity - image")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    btn_berms_application_proceed = driver.find_element(By.XPATH, '//form[@id="form_application"]//div[3]//div[2]//div//div//button')
    driver.execute_script("arguments[0].click();", btn_berms_application_proceed)
    print("Move to next page ")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    input_berms_existing_business_requirements_office = driver.find_element(By.XPATH, '//input[@id="govtOfficeName"]')
    input_berms_existing_business_requirements_office.send_keys("WHY THE OFFICE OF THE WHAT")
    print("Applicant existing business registration - office")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

# try:
#     input_berms_existing_business_requirements_regdate = driver.find_element(By.XPATH, '//input[@id="existingBusinessRegistryDate"]')
#     input_berms_existing_business_requirements_regdate.send_keys("01/10/2025")
#     print("Applicant existing business registration - date")
#     #return 1
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

try:
    input_berms_existing_business_req_regdate = driver.find_element(By.XPATH, '//input[@id="existingBusinessRegistryDate"]')
    input_berms_existing_business_req_regdate_dp = driver.find_element(By.XPATH, '//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[4]//div[2]//div[2]//input[2]')
    driver.execute_script("arguments[0].click();", input_berms_existing_business_req_regdate_dp)
    input_berms_existing_business_req_regdate_dp.clear()
    input_berms_existing_business_req_regdate_dp.send_keys("01/10/2024")
    input_berms_existing_business_req_regdate_dp.send_keys(Keys.TAB)
    time.sleep(2)
    # registration_date = date(2025,1,10)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%m/%d/%Y")) + ";", input_berms_existing_business_req_regdate_dp)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%Y-%m-%d")) + ";", input_berms_existing_business_req_regdate)
    # print("Applicant existing business registration - date ", registration_date.strftime("%Y-%m-%d"), ", ",input_berms_existing_business_req_regdate.get_property("value"),", ", input_berms_existing_business_req_regdate_dp.get_property("value"))
    print("Applicant existing business registration - date ", input_berms_existing_business_req_regdate_dp.get_property("value"), ", ", input_berms_existing_business_req_regdate.get_property("value"))
except NoSuchElementException:
    print("Element Not found! Test Fail")

time.sleep(1)

try:
    input_berms_existing_business_req_regnum = driver.find_element(By.XPATH, '//input[@id="registration_no"]')
    input_berms_existing_business_req_regnum.send_keys("REG1000001")
    print("Applicant existing business registration - registration num")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_existing_business_req_secnum = driver.find_element(By.XPATH, '//input[@id="secPrimaryPurpose"]')
    input_berms_existing_business_req_secnum.send_keys("SEC1000001")
    print("Applicant existing business registration - sec num")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail") nb
    #return 0

time.sleep(1)

try:
    input_berms_existing_business_req_amortamount = driver.find_element(By.XPATH, '//input[@id="authorizedAmount"]')
    input_berms_existing_business_req_amortamount.send_keys("100000")
    print("Applicant existing business registration - amortization amount")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_existing_business_req_subsamount = driver.find_element(By.XPATH, '//input[@id="subscribeAmount"]')
    input_berms_existing_business_req_subsamount.send_keys("100000")
    print("Applicant existing business registration - subscribed amount")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    input_berms_existing_business_req_paidupamount = driver.find_element(By.XPATH, '//input[@id="paidUpAmount"]')
    input_berms_existing_business_req_paidupamount.send_keys("100000")
    print("Applicant existing business registration - paid up amount")
    #return 1
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    btn_berms_rep_proceed = driver.find_element(By.XPATH, '//form[@id="form_application"]//div[4]//div[3]//div//div//button')
    driver.execute_script("arguments[0].click();", btn_berms_rep_proceed)
    print("Move to next page ")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    btn_berms_sholders_add = driver.find_element(By.XPATH, '//button[@id="btnAddStockholder"]')
    driver.execute_script("arguments[0].click();", btn_berms_sholders_add)
    print("Add stockholders - add")
    time.sleep(5)
    try:
        modal_stockholder_add: WebElement = driver.find_element(By.XPATH, '//div[@id="proposedStockholdersModal"]')
        time.sleep(5)
        modal_stockholder_add.find_element(By.XPATH, '//input[@id="proposedStockholderFirstName"]').send_keys("SH FNAME1")
        modal_stockholder_add.find_element(By.XPATH, '//input[@id="proposedStockholderMiddleName"]').send_keys("SH MNAME1")
        modal_stockholder_add.find_element(By.XPATH, '//input[@id="proposedStockholderLastName"]').send_keys("SH LNAME1")
        modal_stockholder_add_nationality = modal_stockholder_add.find_element(By.XPATH, '//select[@id="proposedStockholderNationalityId"]')
        Select(modal_stockholder_add_nationality).select_by_visible_text("FILIPINO")
        modal_stockholder_add.find_element(By.XPATH, '//input[@id="proposedStockholderNumberOfShares"]').send_keys("75")
        modal_stockholder_add.find_element(By.XPATH, '//input[@id="proposedStockholderSubscribeAmt"]').send_keys("500000")
        modal_stockholder_add.find_element(By.XPATH, '//input[@id="proposedStockholderPaidupAmt"]').send_keys("250000")
        time.sleep(5)
        modal_stockholder_add_button = modal_stockholder_add.find_element(By.XPATH, '//button[@id="btnAddStockholder" and @class="btn btn-primary skip-text-change skip-submit-loadin"]')
        driver.execute_script("arguments[0].click();", modal_stockholder_add_button)
        time.sleep(5)
    except NoSuchElementException:
        print("Element Not found! Test Fail")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    radio_berms_othersholders_add = driver.find_element(By.XPATH, '//input[@id="isListedStockholder1"]')
    driver.execute_script("arguments[0].click();", radio_berms_othersholders_add)
    print("Add stockholders - add other")
except NoSuchElementException:
    print("Element Not found! Test Fail")

time.sleep(1)

try:
    btn_berms_othersholders_add = driver.find_element(By.XPATH, '//button[@class="btn btn-secondary btn-sm btn-primary"]')
    driver.execute_script("arguments[0].click();", btn_berms_othersholders_add)
    print("Add stockholders - add other")
    try:
        modal_otherstockholder_add: WebElement = driver.find_element(By.XPATH, '//div[@id="otherStockholdersModal"]')
        time.sleep(5)
        modal_otherstockholder_add_sholder = modal_otherstockholder_add.find_element(By.XPATH, '//select[@id="registeredStockholderName"]')
        Select(modal_otherstockholder_add_sholder).select_by_visible_text("SH LNAME1, SH FNAME1 SH MNAME1")
        modal_otherstockholder_add.find_element(By.XPATH, '//input[@id="stockholderEnterpriseName"]').send_keys("WHY THE OTHER PEZA ENTERPRISE 1")
        modal_otherstockholder_add.find_element(By.XPATH, '//input[@id="stockholderEnterpriseActivity"]').send_keys("WHY THE OTHER PEZA ACTIVITY 1")
        modal_otherstockholder_add.find_element(By.XPATH, '//input[@id="stockholderEnterpriseEquity"]').send_keys("100")
        time.sleep(5)
        modal_otherstockholder_add_button = modal_otherstockholder_add.find_element(By.XPATH, '//button[@id="btnAddStockholderEnterprise"]')
        driver.execute_script("arguments[0].click();", modal_otherstockholder_add_button)
        time.sleep(5)
    except NoSuchElementException:
        print("Element Not found! Test Fail")
except NoSuchElementException:
    print("Element Not found! Test Fail")

time.sleep(1)

try:
    btn_berms_principalofficer_add = driver.find_element(By.XPATH, '//button[@data-bs-target="#principalOfficerModal"]')
    driver.execute_script("arguments[0].click();", btn_berms_principalofficer_add)
    print("Add stockholders - add other")
    try:
        modal_principal_officer_add: WebElement = driver.find_element(By.XPATH, '//div[@id="principalOfficerModal"]')
        time.sleep(5)
        modal_principal_officer_add.find_element(By.XPATH, '//input[@id="officerFirstName"]').send_keys("WHY PRINCIPAL OFFICE FNAME 1")
        modal_principal_officer_add.find_element(By.XPATH, '//input[@id="officerMiddleName"]').send_keys("WHY PRINCIPAL OFFICE MNAME 1")
        modal_principal_officer_add.find_element(By.XPATH, '//input[@id="officerLastName"]').send_keys("WHY PRINCIPAL OFFICE LNAME 1")
        modal_principal_officer_add_nationality = modal_principal_officer_add.find_element(By.XPATH, '//select[@id="officerNationalityId"]')
        Select(modal_principal_officer_add_nationality).select_by_visible_text("NORTHERN IRISH")
        modal_principal_officer_add.find_element(By.XPATH, '//input[@id="officerPosition"]').send_keys("WHY PRINCIPAL OFFICE POSITION")
        time.sleep(5)
        modal_principal_officer_add_button = modal_principal_officer_add.find_element(By.XPATH, '//button[@id="btnAddPrincipalOfficer"]')
        driver.execute_script("arguments[0].click();", modal_principal_officer_add_button)
        time.sleep(5)
    except NoSuchElementException:
        print("Element Not found! Test Fail")
except NoSuchElementException:
    print("Element Not found! Test Fail")

time.sleep(1)

try:
    modal_principal_officer_add_button = modal_principal_officer_add.find_element(By.XPATH, '//button[@id="btnAddPrincipalOfficer"]')
    driver.execute_script("arguments[0].click();", modal_principal_officer_add_button)
except NoSuchElementException:
    print("Element Not found! Test Fail")

time.sleep(1)

try:
    btn_berms_stock_officers_proceed = driver.find_element(By.XPATH, '//form[@id="form_application"]//div[5]//div[5]//div//div//button')
    driver.execute_script("arguments[0].click();", btn_berms_stock_officers_proceed)
    print("Move to next page ")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

# try:
#     btn_berms_man_time_add_man = driver.find_element(By.XPATH, '//button[@id="btnAddManpowerRequirement"]')
#     driver.execute_script("arguments[0].click();", btn_berms_man_time_add_man)
#     print("Move to next page ")
# except NoSuchElementException:
#     print("Element Not found! Test Fail")
#     #return 0

time.sleep(1)

try:
    input_berms_man_time_add_man_year = driver.find_element(By.XPATH, '//select[@name="manpower[0].year"]')
    Select(input_berms_man_time_add_man_year).select_by_visible_text("2025")
    print("Man power requirement - year")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="manpower[0].service_personnel_count"]').send_keys("50")
    print("Man power requirement - service personel")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="manpower[0].indirect_personnel_count"]').send_keys("50")
    print("Man power requirement - service indirect personel")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="manpower[0].admin_personnel_count"]').send_keys("50")
    print("Man power requirement - service admin personel")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    input_berms_man_time_add_man_building_date = driver.find_element(By.XPATH, '//input[@name="project_timetable[0][period_from]"]')
    input_berms_man_time_add_man_building_date_dp = driver.find_element(By.XPATH, '//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[6]//div[5]//div[1]//div[1]//div//div//div[2]//div[1]//input[2]')
    driver.execute_script("arguments[0].click();", input_berms_man_time_add_man_building_date_dp)
    input_berms_man_time_add_man_building_date_dp.clear()
    input_berms_man_time_add_man_building_date_dp.send_keys("01/10/2025")
    input_berms_man_time_add_man_building_date_dp.send_keys(Keys.TAB)
    time.sleep(2)
    # registration_date = date(2025,1,10)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%m/%d/%Y")) + ";", input_berms_existing_business_req_regdate_dp)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%Y-%m-%d")) + ";", input_berms_existing_business_req_regdate)
    # print("Applicant existing business registration - date ", registration_date.strftime("%Y-%m-%d"), ", ",input_berms_existing_business_req_regdate.get_property("value"),", ", input_berms_existing_business_req_regdate_dp.get_property("value"))
    print("Building Construction - from date ", input_berms_man_time_add_man_building_date.get_property("value"), ", ", input_berms_man_time_add_man_building_date_dp.get_property("value"))
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    input_berms_man_time_add_man_building_date = driver.find_element(By.XPATH, '//input[@name="project_timetable[0][period_to]"]')
    input_berms_man_time_add_man_building_date_dp = driver.find_element(By.XPATH, '//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[6]//div[5]//div[1]//div[1]//div//div//div[2]//div[2]//input[2]')
    driver.execute_script("arguments[0].click();", input_berms_man_time_add_man_building_date_dp)
    input_berms_man_time_add_man_building_date_dp.clear()
    input_berms_man_time_add_man_building_date_dp.send_keys("12/10/2025")
    input_berms_man_time_add_man_building_date_dp.send_keys(Keys.TAB)
    time.sleep(2)
    # registration_date = date(2025,1,10)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%m/%d/%Y")) + ";", input_berms_existing_business_req_regdate_dp)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%Y-%m-%d")) + ";", input_berms_existing_business_req_regdate)
    # print("Applicant existing business registration - date ", registration_date.strftime("%Y-%m-%d"), ", ",input_berms_existing_business_req_regdate.get_property("value"),", ", input_berms_existing_business_req_regdate_dp.get_property("value"))
    print("Building Construction - from date ", input_berms_man_time_add_man_building_date.get_property("value"), ", ", input_berms_man_time_add_man_building_date_dp.get_property("value"))
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    input_berms_man_time_add_man_importation_date = driver.find_element(By.XPATH, '//input[@name="project_timetable[1][period_from]"]')
    input_berms_man_time_add_man_importation_date_dp = driver.find_element(By.XPATH, '//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[6]//div[5]//div[1]//div[2]//div//div//div[2]//div[1]//input[2]')
    driver.execute_script("arguments[0].click();", input_berms_man_time_add_man_importation_date_dp)
    input_berms_man_time_add_man_importation_date_dp.clear()
    input_berms_man_time_add_man_importation_date_dp.send_keys("01/11/2025")
    input_berms_man_time_add_man_importation_date_dp.send_keys(Keys.TAB)
    time.sleep(2)
    # registration_date = date(2025,1,10)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%m/%d/%Y")) + ";", input_berms_existing_business_req_regdate_dp)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%Y-%m-%d")) + ";", input_berms_existing_business_req_regdate)
    # print("Applicant existing business registration - date ", registration_date.strftime("%Y-%m-%d"), ", ",input_berms_existing_business_req_regdate.get_property("value"),", ", input_berms_existing_business_req_regdate_dp.get_property("value"))
    print("Importation of machinery - from date ", input_berms_man_time_add_man_importation_date.get_property("value"), ", ", input_berms_man_time_add_man_importation_date_dp.get_property("value"))
except NoSuchElementException:
    print("Element Not found! Test Fail")


try:
    input_berms_man_time_add_man_importation_dateto = driver.find_element(By.XPATH, '//input[@name="project_timetable[1][period_to]"]')
    input_berms_man_time_add_man_importation_dateto_dp = driver.find_element(By.XPATH, '//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[6]//div[5]//div[1]//div[2]//div//div//div[2]//div[2]//input[2]')
    driver.execute_script("arguments[0].click();", input_berms_man_time_add_man_importation_dateto_dp)
    input_berms_man_time_add_man_importation_dateto_dp.clear()
    input_berms_man_time_add_man_importation_dateto_dp.send_keys("12/11/2025")
    input_berms_man_time_add_man_importation_dateto_dp.send_keys(Keys.TAB)
    time.sleep(2)
    # registration_date = date(2025,1,10)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%m/%d/%Y")) + ";", input_berms_existing_business_req_regdate_dp)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%Y-%m-%d")) + ";", input_berms_existing_business_req_regdate)
    # print("Applicant existing business registration - date ", registration_date.strftime("%Y-%m-%d"), ", ",input_berms_existing_business_req_regdate.get_property("value"),", ", input_berms_existing_business_req_regdate_dp.get_property("value"))
    print("Importation of machinery - from date ", input_berms_man_time_add_man_importation_dateto.get_property("value"), ", ", input_berms_man_time_add_man_importation_dateto_dp.get_property("value"))
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    input_berms_man_time_add_man_importation_date = driver.find_element(By.XPATH, '//input[@name="project_timetable[2][period_from]"]')
    input_berms_man_time_add_man_importation_date_dp = driver.find_element(By.XPATH, '//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[6]//div[5]//div[2]//div[1]//div//div//div[2]//div[1]//input[2]')
    driver.execute_script("arguments[0].click();", input_berms_man_time_add_man_importation_date_dp)
    input_berms_man_time_add_man_importation_date_dp.clear()
    input_berms_man_time_add_man_importation_date_dp.send_keys("01/12/2025")
    input_berms_man_time_add_man_importation_date_dp.send_keys(Keys.TAB)
    time.sleep(2)
    # registration_date = date(2025,1,10)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%m/%d/%Y")) + ";", input_berms_existing_business_req_regdate_dp)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%Y-%m-%d")) + ";", input_berms_existing_business_req_regdate)
    # print("Applicant existing business registration - date ", registration_date.strftime("%Y-%m-%d"), ", ",input_berms_existing_business_req_regdate.get_property("value"),", ", input_berms_existing_business_req_regdate_dp.get_property("value"))
    print("Installation of machinery - from date ", input_berms_man_time_add_man_importation_date.get_property("value"), ", ", input_berms_man_time_add_man_importation_date_dp.get_property("value"))
except NoSuchElementException:
    print("Element Not found! Test Fail")


try:
    input_berms_man_time_add_man_importation_dateto = driver.find_element(By.XPATH, '//input[@name="project_timetable[2][period_to]"]')
    input_berms_man_time_add_man_importation_dateto_dp = driver.find_element(By.XPATH, '//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[6]//div[5]//div[2]//div[1]//div//div//div[2]//div[2]//input[2]')
    driver.execute_script("arguments[0].click();", input_berms_man_time_add_man_importation_dateto_dp)
    input_berms_man_time_add_man_importation_dateto_dp.clear()
    input_berms_man_time_add_man_importation_dateto_dp.send_keys("12/12/2025")
    input_berms_man_time_add_man_importation_dateto_dp.send_keys(Keys.TAB)
    time.sleep(2)
    # registration_date = date(2025,1,10)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%m/%d/%Y")) + ";", input_berms_existing_business_req_regdate_dp)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%Y-%m-%d")) + ";", input_berms_existing_business_req_regdate)
    # print("Applicant existing business registration - date ", registration_date.strftime("%Y-%m-%d"), ", ",input_berms_existing_business_req_regdate.get_property("value"),", ", input_berms_existing_business_req_regdate_dp.get_property("value"))
    print("Installation of machinery - from date ", input_berms_man_time_add_man_importation_dateto.get_property("value"), ", ", input_berms_man_time_add_man_importation_dateto_dp.get_property("value"))
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    input_berms_man_time_add_man_importation_date = driver.find_element(By.XPATH, '//input[@name="project_timetable[3][period_from]"]')
    input_berms_man_time_add_man_importation_date_dp = driver.find_element(By.XPATH, '//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[6]//div[5]//div[2]//div[2]//div//div//div[2]//div[1]//input[2]')
    driver.execute_script("arguments[0].click();", input_berms_man_time_add_man_importation_date_dp)
    input_berms_man_time_add_man_importation_date_dp.clear()
    input_berms_man_time_add_man_importation_date_dp.send_keys("01/13/2025")
    input_berms_man_time_add_man_importation_date_dp.send_keys(Keys.TAB)
    time.sleep(2)
    # registration_date = date(2025,1,10)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%m/%d/%Y")) + ";", input_berms_existing_business_req_regdate_dp)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%Y-%m-%d")) + ";", input_berms_existing_business_req_regdate)
    # print("Applicant existing business registration - date ", registration_date.strftime("%Y-%m-%d"), ", ",input_berms_existing_business_req_regdate.get_property("value"),", ", input_berms_existing_business_req_regdate_dp.get_property("value"))
    print("Hiring personel - from date ", input_berms_man_time_add_man_importation_date.get_property("value"), ", ", input_berms_man_time_add_man_importation_date_dp.get_property("value"))
except NoSuchElementException:
    print("Element Not found! Test Fail")


try:
    input_berms_man_time_add_man_importation_dateto = driver.find_element(By.XPATH, '//input[@name="project_timetable[3][period_to]"]')
    input_berms_man_time_add_man_importation_dateto_dp = driver.find_element(By.XPATH, '//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[6]//div[5]//div[2]//div[2]//div//div//div[2]//div[2]//input[2]')
    driver.execute_script("arguments[0].click();", input_berms_man_time_add_man_importation_dateto_dp)
    input_berms_man_time_add_man_importation_dateto_dp.clear()
    input_berms_man_time_add_man_importation_dateto_dp.send_keys("12/13/2025")
    input_berms_man_time_add_man_importation_dateto_dp.send_keys(Keys.TAB)
    time.sleep(2)
    # registration_date = date(2025,1,10)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%m/%d/%Y")) + ";", input_berms_existing_business_req_regdate_dp)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%Y-%m-%d")) + ";", input_berms_existing_business_req_regdate)
    # print("Applicant existing business registration - date ", registration_date.strftime("%Y-%m-%d"), ", ",input_berms_existing_business_req_regdate.get_property("value"),", ", input_berms_existing_business_req_regdate_dp.get_property("value"))
    print("Hiring personel - from date ", input_berms_man_time_add_man_importation_dateto.get_property("value"), ", ", input_berms_man_time_add_man_importation_dateto_dp.get_property("value"))
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    input_berms_man_time_add_man_importation_date = driver.find_element(By.XPATH, '//input[@name="project_timetable[4][period_from]"]')
    input_berms_man_time_add_man_importation_date_dp = driver.find_element(By.XPATH, '//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[6]//div[5]//div[3]//div//div//div//div[2]//div[1]//input[2]')
    driver.execute_script("arguments[0].click();", input_berms_man_time_add_man_importation_date_dp)
    input_berms_man_time_add_man_importation_date_dp.clear()
    input_berms_man_time_add_man_importation_date_dp.send_keys("01/14/2025")
    input_berms_man_time_add_man_importation_date_dp.send_keys(Keys.TAB)
    time.sleep(2)
    # registration_date = date(2025,1,10)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%m/%d/%Y")) + ";", input_berms_existing_business_req_regdate_dp)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%Y-%m-%d")) + ";", input_berms_existing_business_req_regdate)
    # print("Applicant existing business registration - date ", registration_date.strftime("%Y-%m-%d"), ", ",input_berms_existing_business_req_regdate.get_property("value"),", ", input_berms_existing_business_req_regdate_dp.get_property("value"))
    print("Start of commercial - from date ", input_berms_man_time_add_man_importation_date.get_property("value"), ", ", input_berms_man_time_add_man_importation_date_dp.get_property("value"))
except NoSuchElementException:
    print("Element Not found! Test Fail")


try:
    input_berms_man_time_add_man_importation_dateto = driver.find_element(By.XPATH, '//input[@name="project_timetable[4][period_to]"]')
    input_berms_man_time_add_man_importation_dateto_dp = driver.find_element(By.XPATH, '//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[6]//div[5]//div[3]//div//div//div//div[2]//div[2]//input[2]')
    driver.execute_script("arguments[0].click();", input_berms_man_time_add_man_importation_dateto_dp)
    input_berms_man_time_add_man_importation_dateto_dp.clear()
    input_berms_man_time_add_man_importation_dateto_dp.send_keys("12/14/2025")
    input_berms_man_time_add_man_importation_dateto_dp.send_keys(Keys.TAB)
    time.sleep(2)
    # registration_date = date(2025,1,10)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%m/%d/%Y")) + ";", input_berms_existing_business_req_regdate_dp)
    # driver.execute_script("arguments[0].value =" + str(registration_date.strftime("%Y-%m-%d")) + ";", input_berms_existing_business_req_regdate)
    # print("Applicant existing business registration - date ", registration_date.strftime("%Y-%m-%d"), ", ",input_berms_existing_business_req_regdate.get_property("value"),", ", input_berms_existing_business_req_regdate_dp.get_property("value"))
    print("Start of commercial - from date ", input_berms_man_time_add_man_importation_dateto.get_property("value"), ", ", input_berms_man_time_add_man_importation_dateto_dp.get_property("value"))
except NoSuchElementException:
    print("Element Not found! Test Fail")

time.sleep(1)

try:
    btn_berms_stock_officers_proceed = driver.find_element(By.XPATH, '//form[@id="form_application"]//div[6]//div[6]//div//div//button')
    driver.execute_script("arguments[0].click();", btn_berms_stock_officers_proceed)
    print("Move to next page ")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    driver.find_element(By.XPATH, '//textarea[@id="manufacturing_process_service_flow_id"]').send_keys("TEST TEST TEST TEST TEST TEST WHY WHY WHY WHYH")
    print("Manufacturing Process Service Flow - Description")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    input_berms_manufacturing_flow_image = driver.find_element(By.XPATH, '//div[@id="addtlInfo_supporting_documentsDropzone"]')
    file_input = driver.execute_script(JS_DROP_FILE, input_berms_manufacturing_flow_image, 0, 0)
    file_input.send_keys(r"C:\Users\ASCENT\Downloads\viber_image_2024-01-02_15-34-22-566.png")
    print("Manufacturing Process Service Flow - image")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(1)

try:
    btn_berms_manufacture_process_flow_proceed = driver.find_element(By.XPATH, '//form[@id="form_application"]//div[7]//div[6]//div//div//button')
    driver.execute_script("arguments[0].click();", btn_berms_manufacture_process_flow_proceed)
    print("Move to next page ")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    driver.find_element(By.XPATH, '//input[@name="equipments[0].item_descripition"]').send_keys("WHY DESCRIPTION")
    print("Machinery Production Schedule - Description")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@name="equipments[0].item_quantity"]').send_keys("100")
    print("Machinery Production Schedule - Quantity")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@name="equipments[0].item_unit_cost"]').send_keys("100")
    print("Machinery Production Schedule - Cost")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@name="equipments[0].item_source"]').send_keys("WHY THE SOURCE")
    print("Machinery Production Schedule - Source")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@name="addtlInfo.prod_sched_shifts"]').send_keys("3")
    print("Production Schedule - Shifts")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@name="addtlInfo.prod_sched_hr_shifts"]').send_keys("9")
    print("Production Schedule - Hr/Shifts")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@name="addtlInfo.prod_sched_nodays_amonth"]').send_keys("21")
    print("Production Schedule - Hr/Shifts")
except NoSuchElementException:
    print("Element Not found! Test Fail")

time.sleep(1)

try:
    btn_berms_production_schedule_proceed = driver.find_element(By.XPATH, '//form[@id="form_application"]//div[8]//div[6]//div//div//button')
    driver.execute_script("arguments[0].click();", btn_berms_production_schedule_proceed)
    print("Move to next page ")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    driver.find_element(By.XPATH, '//input[@id="floor_area_sqm"]').send_keys("210")
    print("Area utilities and waste disposal - floor area")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@id="total_area_sqm"]').send_keys("1000")
    print("Area utilities and waste disposal - total area")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    Select(driver.find_element(By.XPATH, '//select[@name="ezoneAccreditationAppsBean.zone_id"]')).select_by_visible_text("PASAY CITY ECONOMIC ZONE")
    print("Area utilities and waste disposal - zone")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@id="unitOwnerFirstname"]').send_keys("WHY UNITOWNER FNAME")
    print("Area utilities and waste disposal - unit owner fname")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@id="unitOwnerMiddlename"]').send_keys("WHY UNITOWNER MNAME")
    print("Area utilities and waste disposal - unit owner Mname")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@id="unitOwnerLastname"]').send_keys("WHY UNITOWNER LNAME")
    print("Area utilities and waste disposal - unit owner Lname")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@id="lotOwnerFirstname"]').send_keys("WHY LOTOWNER FNAME")
    print("Area utilities and waste disposal - LOT owner fname")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@id="lotOwnerMiddlename"]').send_keys("WHY LOTOWNER MNAME")
    print("Area utilities and waste disposal - LOT owner mname")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@id="lotOwnerLastname"]').send_keys("WHY LOTOWNER FNAME")
    print("Area utilities and waste disposal - LOT owner lname")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@id="water_yr_req"]').send_keys("10000")
    print("Area utilities and waste disposal - water requirement")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//input[@id="electric_yr_req"]').send_keys("100000")
    print("Area utilities and waste disposal - electricity requirement")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//textarea[@id="waste_disposal_desc"]').send_keys("WHY THE WASTE DISPOSAL PLAN DESCRIPTION WHY THE WASTE DISPOSAL PLAN DESCRIPTION")
    print("Area utilities and waste disposal - waste disposal desc")
except NoSuchElementException:
    print("Element Not found! Test Fail")

try:
    driver.find_element(By.XPATH, '//textarea[@id="waste_disposal_method_desc"]').send_keys("WHY THE WASTE DISPOSAL METHOD DESCRIPTION WHY THE WASTE DISPOSAL METHOD DESCRIPTION")
    print("Area utilities and waste disposal - waste disposal method")
except NoSuchElementException:
    print("Element Not found! Test Fail")

time.sleep(1)

try:
    btn_berms_util_waste_proceed = driver.find_element(By.XPATH, '//form[@id="form_application"]//div[9]//div[7]//div//div//button')
    driver.execute_script("arguments[0].click();", btn_berms_util_waste_proceed)
    print("Move to next page ")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    Select(driver.find_element(By.XPATH, '//select[@name="projectMarketAspectAppsBean.export_percent"]')).select_by_visible_text("50%")
    print("Market Aspect - export")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    Select(driver.find_element(By.XPATH, '//select[@name="projectMarketAspectAppsBean.import_percent"]')).select_by_visible_text("80%")
    print("Market Aspect - import")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    Select(driver.find_element(By.XPATH, '//select[@name="projectMarketAspectAppsBean.country_of_export_market"]')).select_by_visible_text("Azerbaijan")
    print("Market Aspect - county of export")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@id="domestic_clients"]').send_keys("150")
    print("Market Aspect - domestic clients")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@id="export_selling_price"]').send_keys("250")
    print("Market Aspect - export selling price")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@id="projected_volume_sales"]').send_keys("10000000")
    print("Market Aspect - project volume sales")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    Select(driver.find_element(By.XPATH, '//select[@name="projectMarketAspectDetailsAppsBean[0].year_no"]')).select_by_visible_text("2029")
    print("Market Aspect - exporting year")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="projectMarketAspectDetailsAppsBean[0].volume_value"]').send_keys("550")
    print("Market Aspect - exporting volume")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="projectMarketAspectDetailsAppsBean[0].us_value"]').send_keys("450")
    print("Market Aspect - exporting amount")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    Select(driver.find_element(By.XPATH, '//select[@name="projectMarketAspectDetailsAppsBean[1].year_no"]')).select_by_visible_text("2030")
    print("Market Aspect - local sourcing year")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="projectMarketAspectDetailsAppsBean[1].volume_value"]').send_keys("750")
    print("Market Aspect - local sourcing volume")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="projectMarketAspectDetailsAppsBean[1].us_value"]').send_keys("650")
    print("Market Aspect - local sourcing amount")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    btn_berms_market_aspect_proceed = driver.find_element(By.XPATH, '//form[@id="form_application"]//div[10]//div[3]//div//div//button//span')
    driver.execute_script("arguments[0].click();", btn_berms_market_aspect_proceed)
    print("Move to next page ")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="projectCost[building_construction_renovation]"]').send_keys("1000000")
    print("Project Cost and Source of Financing - building construction")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="projectCost[machinery_equipment]"]').send_keys("1","0","00","10","0")
    print("Project Cost and Source of Financing - machinery equipment")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="projectCost[factory_tools]"]').send_keys("1","0","00","20","0")
    print("Project Cost and Source of Financing - factory tools")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="projectCost[transporation]"]').send_keys("1","0","00","30","0")
    print("Project Cost and Source of Financing - transportation")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="projectCost[office_equipment]"]').send_keys("1","0","00","40","0")
    print("Project Cost and Source of Financing - furniture, fixture and equipment")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="projectCost[other_assets]"]').send_keys("1","0","00","50","0")
    print("Project Cost and Source of Financing - other assets")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="projectCost[operating_expenses]"]').send_keys("1","0","00","60","0")
    print("Project Cost and Source of Financing - pre operating expenses")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="projectCost[working_capital]"]').send_keys("1","0","00","70","0")
    print("Project Cost and Source of Financing - working capital")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="fund_source[0][amount]"]').send_keys("0200000")
    print("Project Cost and Source of Financing - equity")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="fund_source[1][amount]"]').send_keys("2","0","00","10","0")
    print("Project Cost and Source of Financing - additional equity")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="fund_source[2][amount]"]').send_keys("2","0","00","20","0")
    print("Project Cost and Source of Financing - advances")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="fund_source[3][amount]"]').send_keys("2","0","00","30","0")
    print("Project Cost and Source of Financing - loans")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    driver.find_element(By.XPATH, '//input[@name="fund_source[4][amount]"]').send_keys("2","0","00","40","0")
    print("Project Cost and Source of Financing - interest rate")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    btn_berms_market_aspect_proceed = driver.find_element(By.XPATH, '//form[@id="form_application"]//div[11]//div[5]//div//div//button//span')
    driver.execute_script("arguments[0].click();", btn_berms_market_aspect_proceed)
    print("Move to next page ")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    driver.find_element(By.XPATH, '//input[@name="supporting_documents[0][filename]"]').send_keys(r"C:\Users\ASCENT\Downloads\eCP Status Workflow_1.pdf")
    print("Project Cost and Source of Financing - SEC")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    input_berms_sup_docs_validity_dp = driver.find_element(By.XPATH, '//html//body//div[2]//div[1]//div//div//div[1]//div[2]//div//div//div[2]//div//div[1]//div//div[2]//form//div[12]//div[2]//div//div//div[1]//div[3]//input[2]')
    driver.execute_script("arguments[0].click();", input_berms_sup_docs_validity_dp)
    input_berms_sup_docs_validity_dp.clear()
    input_berms_sup_docs_validity_dp.send_keys("12/14/2025", Keys.TAB)
    time.sleep(2)
    print("Supporting documents - validity date ",  input_berms_sup_docs_validity_dp.get_property("value"))
except NoSuchElementException:
    print("Element Not found! Test Fail")

time.sleep(5)

try:
    btn_berms_market_aspect_proceed = driver.find_element(By.XPATH, '//form[@id="form_application"]//div[12]//div[3]//div//div//button//span')
    driver.execute_script("arguments[0].click();", btn_berms_market_aspect_proceed)
    print("Move to next page ")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

try:
    input_berms_terms_condition = driver.find_element(By.XPATH, '//input[@id="termsAndCondition"]')
    driver.execute_script("arguments[0].click();", input_berms_terms_condition)
    print("Agree terms and condition")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

driver.save_screenshot(r"c:\Scripts\ScreenShot\0001.png")
time.sleep(320)

try:
    btn_berms_proceed = driver.find_element(By.XPATH, '//button[@id="btnSubmitApplication"]')
    driver.execute_script("arguments[0].click();", btn_berms_proceed)
    print("Submit BERMS Application")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0

time.sleep(5)

try:
    swal_berms_confirmation_to_submit: WebElement = driver.find_element(By.XPATH, '//html//body//div[24]//div')
    swal_berms_confirmation_to_subtmit_btn_ok = swal_berms_confirmation_to_submit.find_element(By.XPATH,'//html//body//div[24]//div//div[6]//button[1]')
    driver.execute_script("arguments[0].click();", swal_berms_confirmation_to_subtmit_btn_ok)
    print("Submit BERMS Application - SWAL Confirmation")
except NoSuchElementException:
    print("Element Not found! Test Fail")
    #return 0


driver.save_screenshot(r"c:\Scripts\ScreenShot\0002.png")
time.sleep(5)


time.sleep(10)

print("Test Completed, please standby")

time.sleep(5)
